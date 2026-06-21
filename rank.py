"""
rank.py — Main entry point. Produces submission.csv and submission.xlsx.

This is the ≤5 minute step that the competition judges will reproduce.
It loads pre-computed semantic scores from artifacts/ and applies
structured feature scoring to rank all 100,000 candidates.

Usage:
    python rank.py --candidates ./data/candidates.jsonl --out ./submission.csv

Prerequisites:
    Run precompute.py first to generate the artifacts/ folder.

Constraints (enforced by competition):
    - ≤ 5 minutes wall-clock time
    - ≤ 16 GB RAM
    - CPU only — no GPU
    - No network access during this step
"""

import argparse
import csv
import json
import logging
import sys
import time
from pathlib import Path

import numpy as np
from tqdm import tqdm

from ranker.loader import stream_candidates
from ranker.reasoning import generate_reasoning
from ranker.scorer import compute_score

# ---------------------------------------------------------------------------
# Setup
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

ARTIFACTS_DIR = Path("artifacts")
TOP_N = 100  # Competition requires exactly 100 ranked candidates


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Rank candidates and produce submission CSV + XLSX."
    )
    parser.add_argument(
        "--candidates",
        type=Path,
        required=True,
        help="Path to candidates.jsonl",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=Path("submission.csv"),
        help="Output CSV path (default: submission.csv)",
    )
    return parser.parse_args()


# ---------------------------------------------------------------------------
# Artifact loading
# ---------------------------------------------------------------------------

def load_artifacts() -> tuple[np.ndarray, dict[str, int]]:
    """Load pre-computed semantic scores and the candidate-ID index.

    Returns:
        Tuple of (scores_array, id_to_index_dict).
        scores_array[i] = MiniLM similarity score for the i-th candidate.
        id_to_index_dict maps candidate_id → array index.
    """
    scores_path = ARTIFACTS_DIR / "semantic_scores.npy"
    index_path  = ARTIFACTS_DIR / "id_to_index.json"

    if not scores_path.exists():
        logger.error("Missing %s — run precompute.py first.", scores_path)
        sys.exit(1)

    if not index_path.exists():
        logger.error("Missing %s — run precompute.py first.", index_path)
        sys.exit(1)

    scores_array = np.load(str(scores_path))
    logger.info("Loaded semantic scores: %d candidates.", len(scores_array))

    with open(index_path, "r", encoding="utf-8") as f:
        id_to_index: dict[str, int] = json.load(f)
    logger.info("Loaded candidate ID index.")

    return scores_array, id_to_index


# ---------------------------------------------------------------------------
# Core ranking pipeline
# ---------------------------------------------------------------------------

def score_all_candidates(
    candidates_path: Path,
    semantic_scores: np.ndarray,
    id_to_index: dict[str, int],
) -> list[dict]:
    """Stream all candidates, score each one, return a list of result dicts.

    Each result dict contains:
        - candidate_id
        - total_score
        - is_honeypot
        - sub_scores
        - candidate   (full dict, needed for reasoning)

    Args:
        candidates_path: Path to candidates.jsonl.
        semantic_scores: Pre-computed MiniLM scores array.
        id_to_index:     Mapping of candidate_id → score array index.

    Returns:
        List of result dicts, one per candidate.
    """
    results: list[dict] = []

    logger.info("Scoring all candidates...")

    for candidate in tqdm(stream_candidates(candidates_path), desc="Scoring", unit="cands"):
        candidate_id = candidate.get("candidate_id", "")

        # Look up the pre-computed semantic score
        index = id_to_index.get(candidate_id)
        if index is not None:
            semantic_score = float(semantic_scores[index])
        else:
            logger.warning("No semantic score for %s — defaulting to 0.", candidate_id)
            semantic_score = 0.0

        # Compute composite score (includes honeypot check)
        result = compute_score(candidate, semantic_score)
        result["candidate_id"] = candidate_id
        result["candidate"]    = candidate  # Keep full profile for reasoning

        results.append(result)

    logger.info("Scored %d candidates.", len(results))
    return results


def select_top_100(results: list[dict]) -> list[dict]:
    """Sort by total_score descending and return exactly 100 candidates.

    Ties are broken by candidate_id ascending (as per competition spec).

    Args:
        results: Full list of scored candidate dicts.

    Returns:
        Sorted top-100 list.
    """
    # Sort: highest score first; ties broken by candidate_id ascending
    sorted_results = sorted(
        results,
        key=lambda r: (-r["total_score"], r["candidate_id"]),
    )

    top_100 = sorted_results[:TOP_N]
    logger.info(
        "Top-100 score range: %.4f (rank 1) → %.4f (rank 100)",
        top_100[0]["total_score"],
        top_100[-1]["total_score"],
    )
    return top_100


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

def write_csv(top_100: list[dict], output_path: Path) -> None:
    """Write the top-100 candidates to a CSV file in competition format.

    Format: candidate_id, rank, score, reasoning
    Scores are non-increasing. Ties break by candidate_id ascending.

    Args:
        top_100:     Sorted list of top-100 result dicts.
        output_path: Path to write the CSV file.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", encoding="utf-8", newline="") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(["candidate_id", "rank", "score", "reasoning"])

        for rank, result in enumerate(top_100, start=1):
            candidate    = result["candidate"]
            candidate_id = result["candidate_id"]
            score        = result["total_score"]
            reasoning    = generate_reasoning(candidate, result, rank)

            writer.writerow([candidate_id, rank, f"{score:.6f}", reasoning])

    logger.info("CSV written → %s", output_path)


def write_xlsx(top_100: list[dict], output_path: Path) -> None:
    """Write the top-100 candidates to a formatted Excel (XLSX) file.

    Formatting:
        - Header row: bold, light-blue background
        - Rank 1-10: green row background
        - Rank 11-50: yellow row background
        - Rank 51-100: white background
        - Score column: formatted as 3 decimal places
        - Auto-width columns

    Args:
        top_100:     Sorted list of top-100 result dicts.
        output_path: Path to write the XLSX file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed — skipping XLSX output. Run: pip install openpyxl")
        return

    workbook  = Workbook()
    worksheet = workbook.active
    worksheet.title = "Ranked Candidates"

    # --- Style definitions ---
    header_font   = Font(bold=True, color="FFFFFF")
    header_fill   = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    top10_fill    = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    top50_fill    = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    default_fill  = PatternFill(fill_type=None)
    wrap_align    = Alignment(wrap_text=True, vertical="top")

    # --- Header row ---
    headers = ["Rank", "Candidate ID", "Score", "Reasoning", "Title", "Company", "Location", "Years Exp"]
    worksheet.append(headers)

    for col_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_index)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- Data rows ---
    for rank, result in enumerate(top_100, start=1):
        candidate = result["candidate"]
        profile   = candidate.get("profile", {})
        reasoning = generate_reasoning(candidate, result, rank)

        row_data = [
            rank,
            result["candidate_id"],
            round(result["total_score"], 4),
            reasoning,
            profile.get("current_title", ""),
            profile.get("current_company", ""),
            profile.get("location", ""),
            profile.get("years_of_experience", 0),
        ]
        worksheet.append(row_data)

        # Apply row colour based on rank tier
        if rank <= 10:
            row_fill = top10_fill
        elif rank <= 50:
            row_fill = top50_fill
        else:
            row_fill = default_fill

        current_row = rank + 1  # offset by header row
        for col_index in range(1, len(headers) + 1):
            cell = worksheet.cell(row=current_row, column=col_index)
            if row_fill.fill_type:
                cell.fill = row_fill
            if col_index == 4:  # Reasoning column — wrap text
                cell.alignment = wrap_align

    # --- Auto-fit column widths (approximate) ---
    column_widths = [8, 16, 10, 60, 30, 25, 20, 12]
    for col_index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(col_index)].width = width

    # Freeze top header row
    worksheet.freeze_panes = "A2"

    xlsx_path = output_path.with_suffix(".xlsx")
    workbook.save(str(xlsx_path))
    logger.info("XLSX written → %s", xlsx_path)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    """Run the full ranking pipeline and write output files."""
    args = parse_args()

    if not args.candidates.exists():
        logger.error("candidates.jsonl not found at: %s", args.candidates)
        sys.exit(1)

    start_time = time.time()

    # Step 1: Load pre-computed artifacts
    semantic_scores, id_to_index = load_artifacts()

    # Step 2: Score every candidate
    all_results = score_all_candidates(args.candidates, semantic_scores, id_to_index)

    # Step 3: Pick the top 100
    top_100 = select_top_100(all_results)

    # Step 4: Write outputs
    write_csv(top_100, args.out)
    write_xlsx(top_100, args.out)

    elapsed = time.time() - start_time
    logger.info("Total runtime: %.1f seconds (%.1f minutes).", elapsed, elapsed / 60)

    if elapsed > 300:
        logger.warning(
            "Runtime exceeded 5-minute competition limit! "
            "Consider reducing BATCH_SIZE in precompute.py or optimising feature extraction."
        )
    else:
        logger.info("✓ Within 5-minute competition budget.")


if __name__ == "__main__":
    main()
